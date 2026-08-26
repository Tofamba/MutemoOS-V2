"""
Unit tests for POST /api/onboarding/bulk-upload — onboards a lawyer plus
their client base from a filled copy of the firm's Client Database Excel
form (fixed layout: B3-B6 lawyer info, row 11 headers, row 12+ client/
matter rows — see backend/main.py::bulk_onboard_from_excel's docstring).

Builds real .xlsx fixtures via openpyxl matching the exact template layout
(same technique as tests/test_bulk_import_clients.py) and calls the
endpoint function directly with a fake DB pool, same convention as this
repo's other backend tests — see tests/test_docx_export.py's docstring for
why (AUTH_ENABLED is False by default, so get_current_user() never
touches `request`).
"""

import asyncio
import io
import uuid

import openpyxl
import pytest
from fastapi import HTTPException

from backend.main import FIRM_ID, bulk_onboard_from_excel


class FakeUploadFile:
    def __init__(self, filename, content: bytes):
        self.filename = filename
        self._content = content

    async def read(self):
        return self._content


def _build_onboarding_xlsx(lawyer, blocks):
    """
    lawyer: {"name", "phone", "email", "role"}
    blocks: list of {"name", "phone", "email", "contact_person", "matters": [text, ...],
                      "client_type" (optional), "beneficial_owner" (optional)}

    Column layout: A=Client Name, B=Phone, C=Email, D=Contact Person
    (companies/entities only), E=Matter description, F=Client Type
    (optional), G=Is the client itself the beneficial owner? (optional).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A3"] = "Your Name:"
    ws["B3"] = lawyer.get("name")
    ws["A4"] = "Your Phone Number:"
    ws["B4"] = lawyer.get("phone")
    ws["A5"] = "Your Email Address:"
    ws["B5"] = lawyer.get("email")
    ws["A6"] = "Your Role:"
    ws["B6"] = lawyer.get("role")
    ws["A11"] = "Client Name (Surname first)"
    ws["B11"] = "Telephone Number"
    ws["C11"] = "Email Address (optional)"
    ws["D11"] = "Contact Person (companies only)"
    ws["E11"] = "Matter (Reference No. — description)"
    ws["F11"] = "Client Type (optional)"
    ws["G11"] = "Client is Beneficial Owner? (optional)"

    row = 12
    for block in blocks:
        matters = block.get("matters") or [None]
        first = True
        for matter_text in matters:
            if first:
                ws[f"A{row}"] = block["name"]
                ws[f"B{row}"] = block.get("phone")
                ws[f"C{row}"] = block.get("email")
                ws[f"D{row}"] = block.get("contact_person")
                ws[f"F{row}"] = block.get("client_type")
                ws[f"G{row}"] = block.get("beneficial_owner")
                first = False
            if matter_text:
                ws[f"E{row}"] = matter_text
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class _FakeTransaction:
    async def __aenter__(self):
        return self

    async def __aexit__(self, *exc):
        return False


class FakeConnection:
    def __init__(self, users=None, clients=None, matters=None, numbering_counters=None, client_compliance=None):
        self.users = users if users is not None else []
        self.clients = clients if clients is not None else []
        self.matters = matters if matters is not None else []
        # Phase 1a consolidation: bulk_onboard_from_excel's commit branch
        # now goes through _create_client_row()/_create_matter_row() ->
        # _next_client_number()/_next_matter_number() -> _allocate_next_seq(),
        # the same atomic numbering_counters machinery
        # tests/test_client_intake.py's fake already needed (client_intake
        # was already on the atomic path before this consolidation; this
        # fake previously had none of this because the old
        # bulk_onboard_from_excel never touched numbering_counters at all
        # -- that was exactly the non-atomic-numbering bug being fixed).
        self.numbering_counters = numbering_counters if numbering_counters is not None else []
        # Compliance-gap fix (2026-08-26): _create_client_row() now also
        # inserts a client_compliance row for every client it creates.
        self.client_compliance = client_compliance if client_compliance is not None else []

    def transaction(self):
        return _FakeTransaction()

    async def fetchval(self, query, *args):
        q = " ".join(query.split())
        if q.startswith("SELECT 1 FROM numbering_counters WHERE firm_id=$1 AND prefix=$2"):
            for c in self.numbering_counters:
                if c["firm_id"] == args[0] and c["prefix"] == args[1]:
                    return 1
            return None
        raise NotImplementedError(f"FakeConnection.fetchval: unhandled query: {q}")

    async def fetchrow(self, query, *args):
        q = " ".join(query.split())
        if q.startswith("SELECT * FROM users WHERE firm_id=$1 AND phone=$2"):
            firm_id, phone = args
            for u in self.users:
                if u["firm_id"] == firm_id and u["phone"] == phone:
                    return dict(u)
            return None
        if q.startswith("SELECT initials FROM users WHERE id=$1"):
            for u in self.users:
                if u["id"] == args[0]:
                    return {"initials": u.get("initials")}
            return None
        if q.startswith("UPDATE numbering_counters SET next_seq = next_seq + 1"):
            firm_id, prefix = args
            for c in self.numbering_counters:
                if c["firm_id"] == firm_id and c["prefix"] == prefix:
                    allocated = c["next_seq"]
                    c["next_seq"] += 1
                    return {"allocated": allocated}
            return None
        if q.startswith("INSERT INTO clients"):
            # _create_client_row() uses fetchrow(...RETURNING *), not
            # execute() -- it needs the inserted row (id, client_number)
            # back. Same generic column/arg zip trick as execute()'s
            # handler below.
            cols = [c.strip() for c in q.split("(", 1)[1].split(")", 1)[0].split(",")]
            row = dict(zip(cols, args))
            self.clients.append(row)
            return dict(row)
        if q.startswith("INSERT INTO matters"):
            # _create_matter_row() likewise uses fetchrow(...RETURNING *).
            cols = [c.strip() for c in q.split("(", 1)[1].split(")", 1)[0].split(",")]
            row = dict(zip(cols, args))
            self.matters.append(row)
            return dict(row)
        raise NotImplementedError(f"FakeConnection.fetchrow: unhandled query: {q}")

    async def fetch(self, query, *args):
        q = " ".join(query.split())
        if q.startswith("SELECT id, full_name, client_number FROM clients WHERE firm_id=$1"):
            firm_id, = args
            return [dict(c) for c in self.clients if c["firm_id"] == firm_id]
        if q.startswith("SELECT initials FROM users WHERE firm_id=$1 AND initials IS NOT NULL"):
            firm_id, = args
            return [{"initials": u["initials"]} for u in self.users if u["firm_id"] == firm_id and u.get("initials")]
        if q.startswith("SELECT client_number FROM clients WHERE firm_id=$1 AND client_number LIKE $2"):
            firm_id, pattern = args
            prefix = pattern[:-1]  # strip trailing '%'
            return [{"client_number": c["client_number"]} for c in self.clients
                    if c["firm_id"] == firm_id and (c.get("client_number") or "").startswith(prefix)]
        if q.startswith("SELECT matter_number FROM matters WHERE firm_id=$1 AND matter_number LIKE $2"):
            firm_id, pattern = args
            prefix = pattern[:-1]  # strip trailing '%'
            return [{"matter_number": m["matter_number"]} for m in self.matters
                    if m["firm_id"] == firm_id and (m.get("matter_number") or "").startswith(prefix)]
        raise NotImplementedError(f"FakeConnection.fetch: unhandled query: {q}")

    async def execute(self, query, *args):
        q = " ".join(query.split())
        if q.startswith("INSERT INTO users"):
            cols = [c.strip() for c in q.split("(", 1)[1].split(")", 1)[0].split(",")]
            self.users.append(dict(zip(cols, args)))
        elif q.startswith("INSERT INTO numbering_counters"):
            firm_id, prefix, seed = args
            if not any(c["firm_id"] == firm_id and c["prefix"] == prefix for c in self.numbering_counters):
                self.numbering_counters.append({"firm_id": firm_id, "prefix": prefix, "next_seq": seed})
        elif q.startswith("INSERT INTO client_compliance"):
            client_id, firm_id, client_is_beneficial_owner = args
            self.client_compliance.append({
                "client_id": client_id, "firm_id": firm_id,
                "client_is_beneficial_owner": client_is_beneficial_owner,
                # Every other column left at its real DB DEFAULT — see
                # _DEFAULT_CLIENT_COMPLIANCE in backend/main.py.
                "identity_verification_status": "Unverified", "risk_rating": "NotAssessed",
                "senior_management_approval_required": False, "enhanced_monitoring_required": False,
                "conflict_check_reviewed": False,
            })
        elif q.startswith("UPDATE users SET initials=$1 WHERE id=$2"):
            initials, user_id = args
            for u in self.users:
                if u["id"] == user_id:
                    u["initials"] = initials
        else:
            raise NotImplementedError(f"FakeConnection.execute: unhandled query: {q}")
        return "OK"


class _FakeAcquireCtx:
    def __init__(self, conn):
        self.conn = conn

    async def __aenter__(self):
        return self.conn

    async def __aexit__(self, *exc):
        return False


class FakePool:
    def __init__(self, users=None, clients=None, matters=None, numbering_counters=None, client_compliance=None):
        self.conn = FakeConnection(users, clients, matters, numbering_counters, client_compliance)

    def acquire(self):
        return _FakeAcquireCtx(self.conn)


LAWYER = {"name": "Tendai Moyo", "phone": "+263771112222", "email": "tendai@sm.co.zw", "role": "Associate"}


# ── clean single-client single-matter ───────────────────────────────────────

def test_clean_single_client_single_matter(monkeypatch):
    import backend.main as m
    pool = FakePool()
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [{
        "name": "Huang Li Qiang", "phone": "+263771234567", "email": "huang@example.com",
        "matters": ["HC 1234/26 — Debt collection"],
    }])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=True))

    assert result["lawyer"]["action"] == "created"
    assert result["lawyer"]["initials"] == "TM"  # generated from "Tendai Moyo"
    assert len(result["clients"]["created"]) == 1
    assert result["clients"]["created"][0]["name"] == "Huang Li Qiang"
    assert result["clients"]["created"][0]["client_number"] == "TM-001"
    assert len(result["matters"]["created"]) == 1
    assert result["matters"]["created"][0]["name"] == "HC 1234/26 — Debt collection"
    assert result["matters"]["created"][0]["matter_number"] == "TM-001-01"
    # Opportunistic, best-effort classification from the matter description.
    assert result["matters"]["created"][0]["practice_area"] == "Debt Collection"
    assert result["clients"]["review"] == []

    assert len(pool.conn.users) == 1
    assert len(pool.conn.clients) == 1
    assert len(pool.conn.matters) == 1
    assert pool.conn.matters[0]["client_id"] == pool.conn.clients[0]["id"]
    assert pool.conn.clients[0]["client_number"] == "TM-001"
    assert pool.conn.matters[0]["matter_number"] == "TM-001-01"
    assert pool.conn.matters[0]["practice_area"] == "Debt Collection"


# ── opportunistic practice_area classification ───────────────────────────

def test_ambiguous_or_unrecognized_matter_text_leaves_practice_area_null(monkeypatch):
    """Best-effort only — never guesses, and never blocks the upload."""
    import backend.main as m
    pool = FakePool()
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [{
        "name": "Huang Li Qiang", "phone": "+263771234567", "email": None,
        "matters": [
            "Mukweva and Paswa Civil — Debt collection/fraud",  # ambiguous: Debt Collection vs Criminal
            "Moyo v Dube — Hearing on Rule Nisi",                # no keyword match at all
        ],
    }])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=True))

    assert len(result["matters"]["created"]) == 2
    assert all(mm["practice_area"] is None for mm in result["matters"]["created"])
    assert all(row["practice_area"] is None for row in pool.conn.matters)


# ── multi-matter client block ────────────────────────────────────────────

def test_multi_matter_client_block(monkeypatch):
    import backend.main as m
    pool = FakePool()
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [{
        "name": "Huang Li Qiang", "phone": "+263771234567", "email": "huang@example.com",
        "matters": [
            "Mukweva and Paswa Civil — Debt collection/fraud, replicated to special plea",
            "Mukweva Criminal — Criminal fraud, following up constitutional application",
            "Kuvimba Mining Corporation — Debt collection, settlement in place",
        ],
    }])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=True))

    assert len(result["clients"]["created"]) == 1  # still one client
    assert len(result["matters"]["created"]) == 3   # three matters, same client
    client_id = result["clients"]["created"][0]["client_id"]
    assert all(mm["row"] == result["clients"]["created"][0]["row"] for mm in result["matters"]["created"])
    assert len(pool.conn.matters) == 3
    assert all(str(row["client_id"]) == client_id for row in pool.conn.matters)

    # Matter numbers are sequential within the one client, in upload order.
    assert [mm["matter_number"] for mm in result["matters"]["created"]] == [
        "TM-001-01", "TM-001-02", "TM-001-03",
    ]


# ── existing-user-phone match (no duplicate) ────────────────────────────

def test_existing_user_phone_matches_no_duplicate(monkeypatch):
    import backend.main as m
    existing_users = [{
        "id": "existing-user-id", "firm_id": FIRM_ID, "phone": LAWYER["phone"],
        "display_name": "Tendai T. Moyo", "role": "partner", "email": "t.moyo@sm.co.zw",
        "initials": "TM",
    }]
    pool = FakePool(users=existing_users)
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [{
        "name": "Vengesai Enterprises", "phone": "+263772223333", "email": None, "matters": ["HC 99/26 — Contract dispute"],
    }])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=True))

    assert result["lawyer"]["action"] == "matched"
    assert result["lawyer"]["user_id"] == "existing-user-id"
    assert result["lawyer"]["display_name"] == "Tendai T. Moyo"  # existing record, not overwritten
    assert result["lawyer"]["initials"] == "TM"  # existing initials reused, not regenerated
    assert len(pool.conn.users) == 1  # no duplicate created
    assert pool.conn.matters[0]["created_by"] == "existing-user-id"
    assert result["clients"]["created"][0]["client_number"] == "TM-001"


# ── matched existing client continues its own matter sequence ───────────

def test_matched_existing_client_matter_numbering_continues_its_sequence(monkeypatch):
    import backend.main as m
    client_id = str(uuid.uuid4())
    existing_clients = [
        {"id": client_id, "firm_id": FIRM_ID, "full_name": "Huang Li Qiang", "client_number": "TM-005"},
    ]
    existing_matters = [
        {"id": "mt1", "firm_id": FIRM_ID, "client_id": client_id, "matter_number": "TM-005-01"},
    ]
    pool = FakePool(clients=existing_clients, matters=existing_matters)
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [{
        "name": "Huang Li Qiang", "phone": "+263771234567", "email": None,
        "matters": ["HC 2/26 — Second matter for existing client"],
    }])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=True))

    assert result["clients"]["matched"][0]["matched_client_number"] == "TM-005"
    assert result["matters"]["created"][0]["matter_number"] == "TM-005-02"


# ── preview mode computes but never persists numbers ─────────────────────

def test_preview_shows_numbers_without_persisting_or_reserving_them(monkeypatch):
    import backend.main as m
    pool = FakePool()
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [{
        "name": "Huang Li Qiang", "phone": "+263771234567", "email": None, "matters": ["HC 1/26 — Test"],
    }])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=False))

    assert result["lawyer"]["initials"] == "TM"
    assert result["clients"]["created"][0]["client_number"] == "TM-001"
    assert result["matters"]["created"][0]["matter_number"] == "TM-001-01"
    # Preview must not write anything, including the lawyer's initials.
    assert pool.conn.users == []
    assert pool.conn.clients == []
    assert pool.conn.matters == []


# ── ambiguous client name lands in review, not auto-merged ──────────────

def test_ambiguous_client_name_lands_in_review_not_auto_merged(monkeypatch):
    import backend.main as m
    existing_clients = [
        {"id": "c1", "firm_id": FIRM_ID, "full_name": "John Moyo", "client_number": "TM-001"},
        {"id": "c2", "firm_id": FIRM_ID, "full_name": "Jon Moyo", "client_number": "TM-002"},
    ]
    pool = FakePool(clients=existing_clients)
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [{
        "name": "J. Moyo", "phone": "+263773334444", "email": None, "matters": ["HC 55/26 — Eviction"],
    }])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=True))

    assert result["clients"]["created"] == []
    assert result["clients"]["matched"] == []
    assert len(result["clients"]["review"]) == 1
    review_entry = result["clients"]["review"][0]
    assert review_entry["name"] == "J. Moyo"
    assert {c["id"] for c in review_entry["candidates"]} == {"c1", "c2"}

    # No new client was created for the ambiguous name.
    assert len(pool.conn.clients) == 2

    # The matter still gets created (nothing silently dropped) but with no
    # client_id — the legacy fallback state — and shows up in the
    # unlinked/pending-review bucket, not the normal "created" list.
    assert result["matters"]["created"] == []
    assert len(result["matters"]["created_unlinked_pending_review"]) == 1
    assert len(pool.conn.matters) == 1
    assert pool.conn.matters[0]["client_id"] is None
    assert pool.conn.matters[0]["client_name"] == "J. Moyo"


# ── corporate/entity name parsed as-is (not surname-flipped) ────────────

def test_corporate_entity_name_stored_as_is(monkeypatch):
    import backend.main as m
    pool = FakePool()
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [{
        "name": "Vengesai Enterprises", "phone": "+263774445555", "email": "info@vengesai.co.zw",
        "matters": ["HC 200/26 — Commercial lease dispute"],
    }])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=True))

    assert result["clients"]["created"][0]["name"] == "Vengesai Enterprises"
    assert pool.conn.clients[0]["full_name"] == "Vengesai Enterprises"


# ── contact_person: corporate client sets it, individual leaves it blank ──

def test_corporate_client_with_contact_person(monkeypatch):
    import backend.main as m
    pool = FakePool()
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [{
        "name": "Vengesai Enterprises", "phone": "+263774445555", "email": "info@vengesai.co.zw",
        "contact_person": "Jane Muzenda (Company Secretary)",
        "matters": ["HC 200/26 — Commercial lease dispute"],
    }])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=True))

    assert result["clients"]["created"][0]["contact_person"] == "Jane Muzenda (Company Secretary)"
    assert pool.conn.clients[0]["contact_person"] == "Jane Muzenda (Company Secretary)"


def test_individual_client_leaves_contact_person_blank(monkeypatch):
    """The common case — no Contact Person column entry for a person."""
    import backend.main as m
    pool = FakePool()
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [{
        "name": "Huang Li Qiang", "phone": "+263771234567", "email": "huang@example.com",
        "matters": ["HC 1234/26 — Debt collection"],
    }])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=True))

    assert result["clients"]["created"][0]["contact_person"] is None
    assert pool.conn.clients[0]["contact_person"] is None


# ── preview (commit=False) writes nothing ────────────────────────────────

def test_preview_does_not_write(monkeypatch):
    import backend.main as m
    pool = FakePool()
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [{
        "name": "Huang Li Qiang", "phone": "+263771234567", "email": None, "matters": ["HC 1/26 — Test"],
    }])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=False))

    assert result["committed"] is False
    assert result["lawyer"]["action"] == "would_create"
    assert len(result["clients"]["created"]) == 1
    assert len(result["matters"]["created"]) == 1
    # Nothing actually written to the fake DB.
    assert pool.conn.users == []
    assert pool.conn.clients == []
    assert pool.conn.matters == []


# ── header validation ────────────────────────────────────────────────────

def test_missing_lawyer_phone_rejected(monkeypatch):
    import backend.main as m
    pool = FakePool()
    monkeypatch.setattr(m, "_db_pool", pool)

    bad_lawyer = dict(LAWYER)
    bad_lawyer["phone"] = None
    content = _build_onboarding_xlsx(bad_lawyer, [{"name": "Huang Li Qiang", "matters": ["HC 1/26 — Test"]}])

    with pytest.raises(HTTPException) as exc_info:
        asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=False))
    assert exc_info.value.status_code == 422
    assert "phone" in exc_info.value.detail.lower()


def test_rejects_non_xlsx_file():
    with pytest.raises(HTTPException) as exc_info:
        asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.docx", b"not really xlsx"), commit=False))
    assert exc_info.value.status_code == 422


# ── client_compliance row created at bulk-onboard time (2026-08-26) ────────
# Confirmed in Phase 0's field-by-field diff: bulk-migrated clients
# previously got no client_compliance row at all, unlike clients created
# via the normal single-client UI flow. _create_client_row() now creates
# one for every client it creates, regardless of caller.

def test_blank_compliance_columns_still_create_a_not_yet_assessed_row(monkeypatch):
    """Columns F/G (Client Type, Beneficial Owner) are optional — leaving
    both blank, the normal/expected case, must still produce a real
    client_compliance row in the default "not yet assessed" state (the
    same state _DEFAULT_CLIENT_COMPLIANCE encodes), not no row at all."""
    import backend.main as m
    pool = FakePool()
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [{
        "name": "Huang Li Qiang", "phone": "+263771234567", "email": "huang@example.com",
        "matters": ["HC 1234/26 — Debt collection"],
        # client_type / beneficial_owner deliberately omitted — blank cells
    }])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=True))

    assert result["clients"]["created"][0]["client_number"] == "TM-001"
    assert len(pool.conn.client_compliance) == 1
    row = pool.conn.client_compliance[0]
    assert row["client_id"] == pool.conn.clients[0]["id"]
    assert row["client_is_beneficial_owner"] is None
    assert row["identity_verification_status"] == "Unverified"
    assert row["risk_rating"] == "NotAssessed"
    # client_type wasn't supplied either — the clients row itself stays
    # unset, same as every other creation path today.
    assert pool.conn.clients[0].get("client_type") is None


def test_filled_compliance_columns_populate_client_type_and_beneficial_owner(monkeypatch):
    """When a lawyer/secretary does know these off-hand, both columns
    parse correctly and land in the right table — client_type on
    clients, client_is_beneficial_owner on client_compliance."""
    import backend.main as m
    pool = FakePool()
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [{
        "name": "Zenith Holdings (Pvt) Ltd", "phone": "+263771234568", "email": "zenith@example.com",
        "contact_person": "Farai Muzenda",
        "matters": ["HC 2000/26 — Commercial dispute"],
        "client_type": "Company", "beneficial_owner": "No",
    }])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=True))

    assert result["clients"]["created"][0]["client_number"] == "TM-001"
    assert pool.conn.clients[0]["client_type"] == "Company"
    assert len(pool.conn.client_compliance) == 1
    assert pool.conn.client_compliance[0]["client_is_beneficial_owner"] == "No"


def test_compliance_columns_are_case_insensitive_and_tolerate_a_typo(monkeypatch):
    """Client Type matches CLIENT_TYPES case-insensitively; an
    unrecognized value in either column is treated as blank rather than
    blocking the upload — same "opportunistic, best-effort, never
    blocks" discipline as practice_area classification just below it."""
    import backend.main as m
    pool = FakePool()
    monkeypatch.setattr(m, "_db_pool", pool)

    content = _build_onboarding_xlsx(LAWYER, [
        {
            "name": "Tariro Chikafu", "phone": "+263771234569",
            "matters": ["HC 3000/26 — Estate matter"],
            "client_type": "individual",  # lowercase — must still match "Individual"
            "beneficial_owner": "YES",    # uppercase — must still match "Yes"
        },
        {
            "name": "Some Trust", "phone": "+263771234570",
            "matters": ["HC 3001/26 — Trust matter"],
            "client_type": "Compnay",  # typo — not a real CLIENT_TYPES value
            "beneficial_owner": "Maybe",  # not Yes/No
        },
    ])
    result = asyncio.run(bulk_onboard_from_excel(None, FakeUploadFile("form.xlsx", content), commit=True))

    assert len(pool.conn.clients) == 2
    by_name = {c["full_name"]: c for c in pool.conn.clients}
    assert by_name["Tariro Chikafu"]["client_type"] == "Individual"
    assert by_name["Some Trust"]["client_type"] is None  # typo -> treated as blank, not an error

    compliance_by_client = {row["client_id"]: row for row in pool.conn.client_compliance}
    assert compliance_by_client[by_name["Tariro Chikafu"]["id"]]["client_is_beneficial_owner"] == "Yes"
    assert compliance_by_client[by_name["Some Trust"]["id"]]["client_is_beneficial_owner"] is None
    # Neither row was rejected — result still reports both matters created.
    assert len(result["matters"]["created"]) == 2
