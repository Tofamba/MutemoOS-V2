#!/usr/bin/env python3
"""
Client Intake Template Generator
====================================================================
Generates the blank Excel form a firm fills in to bulk-onboard a lawyer
plus their existing client base via POST /api/onboarding/bulk-upload
(backend/main.py::bulk_onboard_from_excel). Not run at app startup or in
CI -- run manually whenever the layout changes:

    python scripts/generate_client_intake_template.py

Writes frontend/MutemoDesk_Client_Intake_Template.xlsx, alongside
MutemoDesk_Matter_Import_Template.xlsx (the sibling template for
bulk_import_matters, served today by GET /api/matters/template-excel) --
no equivalent download route exists yet for this one; wiring one up the
same way is a natural follow-up, not done here since it wasn't asked for.
Uses openpyxl, already a project dependency (this same endpoint reads
with it).

LAYOUT -- must match bulk_onboard_from_excel()'s parser exactly (see
that function's own docstring/comments in backend/main.py for the
parsing side of this contract; this script is the writing side of the
same contract, kept in sync by hand since the two can't share code
across a script/endpoint boundary):
    B3/B4/B5/B6  -- lawyer Name / Phone / Email / Role ("Partner" or
                    "Associate", case-insensitive on read)
    Row 11       -- column headers. Purely for the human filling this
                    in -- the parser is entirely position-based and
                    never reads row 11 itself.
    Row 12+      -- one row per matter. Columns A/B/C/D/F/G are read
                    only on a client's first row and expected blank on
                    that same client's subsequent matter rows; column D
                    (Contact Person) is companies/entities only, blank
                    for individuals. Column E is one matter's free-text
                    "Reference/Case No. -- description", stored as-is.

    Columns F (Client Type) and G (Is the client itself the beneficial
    owner?) are OPTIONAL compliance fields (2026-08-26) -- populated
    directly into the new client's client_compliance row when filled
    in (backend/main.py::_create_client_row()), left in their default
    "not yet assessed" state otherwise. Leave blank if unsure: a blank
    cell is the normal, expected input for both, never an error, and
    nothing else about compliance (PEP status, source of wealth, senior
    management approval, etc.) belongs in this template -- those
    genuinely require in-person verification, not a bulk-migration
    spreadsheet cell, the same reasoning that kept the RBZ compliance
    export's own source form simple for non-technical staff.

Unlike bulk_import_matters' template (a flexible header-alias parser
that automatically skips one EXAMPLE row) this form's cell positions
ARE the contract, so no example row is written at row 12 here -- that
would be parsed as a real client.
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

OUT_PATH = os.path.join(
    os.path.dirname(__file__), "..", "frontend", "MutemoDesk_Client_Intake_Template.xlsx"
)

# Must match backend/main.py's CLIENT_TYPES exactly -- this is the same
# picklist update_client()/update_client_compliance() validate against.
CLIENT_TYPES = [
    "Individual", "Company", "Partnership", "Trust", "Estate",
    "NonProfit", "Government", "Other",
]

# Generous headroom for a firm's whole existing client base in one
# upload -- optional validation, so a blank cell (the expected default
# for nearly every row) is always valid regardless of how far down it goes.
DATA_VALIDATION_ROWS = 500

HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LABEL_FONT = Font(bold=True)
NOTE_FONT = Font(italic=True, size=9, color="666666")


def build_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Client Intake"

    ws["A1"] = "MutemoOS — Client & Matter Bulk Onboarding"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        "Fill in your details below, then one row per client starting at row 12 "
        "(one extra row per additional matter for the same client, columns A-D "
        "left blank on those extra rows)."
    )
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:G2")
    ws["A2"].alignment = Alignment(wrap_text=True)

    # ── Lawyer info block (B3-B6) ──────────────────────────────────────
    ws["A3"] = "Your Name:"
    ws["A3"].font = LABEL_FONT
    ws["A4"] = "Your Phone Number:"
    ws["A4"].font = LABEL_FONT
    ws["A5"] = "Your Email Address:"
    ws["A5"].font = LABEL_FONT
    ws["A6"] = "Your Role:"
    ws["A6"].font = LABEL_FONT
    ws["C6"] = "(Partner or Associate)"
    ws["C6"].font = NOTE_FONT

    role_dv = DataValidation(
        type="list", formula1='"Partner,Associate"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid Role",
        error='Enter "Partner" or "Associate".',
    )
    ws.add_data_validation(role_dv)
    role_dv.add("B6")

    # ── Row 11 headers ──────────────────────────────────────────────────
    headers = {
        "A11": "Client Name (Surname first)",
        "B11": "Telephone Number",
        "C11": "Email Address (optional)",
        "D11": "Contact Person (companies only)",
        "E11": "Matter (Reference No. — description)",
        "F11": "Client Type (optional — leave blank if unsure)",
        "G11": "Client is Beneficial Owner? (optional — leave blank if unsure)",
    }
    for coord, text in headers.items():
        cell = ws[coord]
        cell.value = text
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 20
    ws.row_dimensions[11].height = 32

    # ── Optional-column dropdowns, rows 12..12+DATA_VALIDATION_ROWS ─────
    last_row = 11 + DATA_VALIDATION_ROWS
    client_type_dv = DataValidation(
        type="list", formula1=f'"{",".join(CLIENT_TYPES)}"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid Client Type",
        error="Pick one of the listed types, or leave the cell blank if unsure.",
    )
    ws.add_data_validation(client_type_dv)
    client_type_dv.add(f"F12:F{last_row}")

    beneficial_owner_dv = DataValidation(
        type="list", formula1='"Yes,No"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid value",
        error='Enter "Yes", "No", or leave the cell blank if unsure.',
    )
    ws.add_data_validation(beneficial_owner_dv)
    beneficial_owner_dv.add(f"G12:G{last_row}")

    ws.freeze_panes = "A12"
    return wb


if __name__ == "__main__":
    wb = build_workbook()
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
